import datetime
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver.common.by import By

from TestEnvironment.GlobalElementCount.ElementCount import ElementCountCls
from TestEnvironment.GlobalFormFill.FormFill import FormFillCls
from chrome.LatestChrome import ChromeCls
from TestEnvironment.GlobalClassMethods.MasterDataExcelReader import DataReadMaster
from TestEnvironment.GlobalElementAction.ElementAction import ElementActionCls
from TestEnvironment.GlobalElementAction.SafeToElementAction import SafeToElementActionCls, SafeToVerify
from TestEnvironment.GlobalElementPresent.ElementPresent import ElementPresentCls
from TestEnvironment.GlobalLoader.Loader import LoaderCls

@allure.step("Entering username ")
def enter_username(username):
    driver.find_element(By.ID, "un").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
    driver.find_element(By.ID, "pw").send_keys(password)

@pytest.fixture()
def test_setup():
    global driver, TestResult, TestResultStatus, path, FundNameList, FundNameListAfterRemove, ct, Exe, D1, D2, d1, d2, DollarDate

    TestResult = []
    TestResultStatus = []
    TestFailStatus = []
    FailStatus = "Pass"
    Exe = "Yes"

    path = DataReadMaster.Path + DataReadMaster.GlobalData("test_Regression_Properties_Admi", "ParentDirectory") +DataReadMaster.GlobalData("test_Regression_Properties_Admi", "Directory") + DataReadMaster.GlobalData("test_Regression_Properties_Admi", "SubDirectory")
    FundNameList = []
    FundNameListAfterRemove = []

    ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
    ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

    today = datetime.date.today()
    D1 = today.strftime("%Y-%m-%d")
    d1 = D1
    DollarDate = datetime.datetime.strptime(d1, '%Y-%m-%d')
    DollarDate = "$" + DollarDate.date().__str__() + "$"
    d1 = datetime.datetime.strptime(D1, "%Y-%m-%d")

    Exe = DataReadMaster.GlobalData("test_Regression_Properties_Admi", "Execution")

    # --------Login to the application-----------------------
    if Exe == "Yes":
        ChromeCls.ChromeMeth()
        driver = webdriver.Chrome(executable_path=ChromeCls.NewChromePath1ChrCls)
        driver.implicitly_wait(10)
        driver.maximize_window()
        driver.get(DataReadMaster.GlobalData("GlobalData", "URL1"))
        enter_username(DataReadMaster.GlobalData("GlobalData", "AdminUsername"))
        enter_password(DataReadMaster.GlobalData("GlobalData", "AdminPassword"))
        driver.find_element(By.XPATH, DataReadMaster.GlobalData("GlobalData", "LoginSubmit")).click()

    yield
    if Exe == "Yes":
        class PDF(FPDF):
            def header(self):
                self.image(path + 'EmailReportContent/Logo.png', 10, 8, 33)
                self.set_font('Arial', 'B', 15)
                self.cell(73)
                self.set_text_color(0, 0, 0)
                self.cell(35, 10, ' Test Report ', 1, 1, 'B')
                self.set_font('Arial', 'I', 10)
                self.cell(150)
                self.cell(30, 10, ctReportHeader, 0, 0, 'C')
                self.ln(20)

            def footer(self):
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.set_text_color(0, 0, 0)
                self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

        pdf = PDF()
        pdf.alias_nb_pages()
        pdf.add_page()
        pdf.set_font('Times', '', 12)
        pdf.cell(0, 10, "Test Case Name:  " + DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PDFTestName"), 0, 1)
        pdf.multi_cell(0, 10,
                       "Description:  " + DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PDFDescription"), 0, 1)
        for i1 in range(len(TestResult)):
            pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(0, 0, 0)
            if (TestResultStatus[i1] == "Fail"):
                # print("Fill Red color")
                pdf.set_text_color(255, 0, 0)
                TestFailStatus.append("Fail")
            TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
            pdf.multi_cell(0, 7, str(i1 + 1) + ")  " + TestName1, 0, 1, fill=True)
            TestFailStatus.append("Pass")
        pdf.output(DataReadMaster.GlobalData("test_Regression_Properties_Admi", "TestName") + "_" + ct + ".pdf")

        # -----------To check if any failed Test case present-------------------
        for io in range(len(TestResult)):
            if TestFailStatus[io] == "Fail":
                FailStatus = "Fail"
        # ---------------------------------------------------------------------

        # -----------To add test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        print()
        check = DataReadMaster.GlobalData("test_Regression_Properties_Admi", "TestName")
        PdfName = DataReadMaster.GlobalData("test_Regression_Properties_Admi", "TestName") + "_" + ct + ".pdf"
        checkcount = 0

        for i in range(1, 100):
            if sheet.cell(i, 1).value == None:
                if checkcount == 0:
                    sheet.cell(row=i, column=1).value = check
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = DataReadMaster.GlobalData("test_Regression_Properties_Admi",
                                                                                  "TestDirectoryName")
                    sheet.cell(row=i, column=4).value = DataReadMaster.GlobalData("test_Regression_Properties_Admi",
                                                                                  "PDFDescription")
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
                wb.save(loc)
                break
            else:
                if sheet.cell(i, 1).value == check:
                    if checkcount == 0:
                        sheet.cell(row=i, column=2).value = PdfName
                        sheet.cell(row=i, column=3).value = DataReadMaster.GlobalData("test_Regression_Properties_Admi",
                                                                                      "TestDirectoryName")
                        sheet.cell(row=i, column=4).value = DataReadMaster.GlobalData("test_Regression_Properties_Admi",
                                                                                      "PDFDescription")
                        sheet.cell(row=i, column=5).value = FailStatus
                        checkcount = 1
        # ----------------------------------------------------------------------------

        # ---------------------To add Test name in Execution sheet--------------------
        ExcelFileName1 = "Execution"
        loc1 = (path + 'Executiondir/' + ExcelFileName1 + '.xlsx')
        wb1 = openpyxl.load_workbook(loc1)
        sheet1 = wb1.active
        checkcount1 = 0

        for ii1 in range(1, 100):
            if sheet1.cell(ii1, 1).value == None:
                if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
                wb1.save(loc1)
                break
            else:
                if sheet1.cell(ii1, 1).value == check:
                    if checkcount1 == 0:
                        sheet1.cell(row=ii1, column=1).value = check
                        checkcount1 = 1
        # -----------------------------------------------------------------------------
        #driver.quit()

@pytest.mark.smoke
def test_AllModules(test_setup):
    if Exe == "Yes":
        try:
            # --------------------Checking Properties Page--------------
            LoaderCls.LoaderMeth(driver)
            PageName = DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PageName")
            driver.find_element(By.XPATH, DataReadMaster.GlobalData("test_Regression_Properties_Admi", "RecordsTab")).click()
            LoaderCls.LoaderMeth(driver)
            driver.find_element(By.XPATH, DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PageLink")).click()
            LoaderCls.LoaderMeth(driver)
            try:
                driver.find_element(By.XPATH,DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PageText")).is_displayed()
                TestResult.append(PageName + " module opened successfully")
                TestResultStatus.append("Pass")
            except Exception as e1:
                print(e1)
                TestResult.append(PageName + " module was not able to open ")
                TestResultStatus.append("Fail")

            # -------------------Add New Property scenario------------------------------
            ElementVerify = "Add New Property button"
            PageName = DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PageName")
            MdataSheetTab = "test_Regression_Properties_Admi"
            MdataSheetItem = "AddNewPropertyButton"
            MdataSheetItem2 = "AddNewPropertyButtonText"
            ElementExpected = "Add New Property"

            #-------To find property in the listing------------
            FieldIDExcel = DataReadMaster.GlobalDataForm(MdataSheetTab, "NameOfProperty$Str")
            driver.find_element(By.XPATH,
                                DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PropertySearch")).send_keys(str(FieldIDExcel[1]))
            driver.find_element(By.XPATH,
                                DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PropertySearchButton")).click()
            LoaderCls.LoaderMeth(driver)
            SearchedEle=driver.find_element(By.XPATH,
                                DataReadMaster.GlobalData("test_Regression_Properties_Admi", "SearchedEle")).text
            substr = "- "
            x = SearchedEle.split(substr)
            SearchedEle=x[1]
            time.sleep(5)
            if SearchedEle==str(FieldIDExcel[1]):
                print("Property present so going inside for more details")
                try:
                    TestResult.append("Search property action worked for property  [ " + SearchedEle + " ]")
                    TestResultStatus.append("Pass")
                except Exception:
                    print("Something went wrong while searching property")
            else:
                print("No Property found so adding new property")
                #----------Add new property-------------
                ItemList = ["NameOfProperty$Str", "Address$Str","City$Str","BedroomCount$Str","BathroomCount$Str","State/Province$Str","SquareFeet$Str","Zip/PostalCode$Str","Country$Str","Description$Str","Status$Clk","Superintendent$Drp","Region$Drp","Resident$Drp"]
                FormFillCls.FormFillMeth(driver,ItemList, MdataSheetTab, MdataSheetItem, MdataSheetItem2,
                                                   ElementExpected,
                                                   ElementVerify,
                                                   PageName, TestResult, TestResultStatus)
                #driver.find_element(By.XPATH, "//button[contains(text(),'Cancel')]").click()
                try:
                    driver.find_element(By.XPATH, "//button[contains(text(),'Submit')]").click()
                    TestResult.append("New property [ " + DataReadMaster.GlobalDataForm(MdataSheetTab,"NameOfProperty$Str") + " ] was added successfully")
                    TestResultStatus.append("Pass")
                except Exception:
                    print("Something went wrong while adding property")
                JSClick = driver.find_element(By.XPATH,
                                              DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PageLink"))
                driver.execute_script("arguments[0].click();", JSClick)

            #----------------------Property Details-------------------------
            MdataSheetTab = "test_Regression_Properties_Admi"
            MdataSheetItem = "SafeToPropertyLinkText"
            SafeToClick = SafeToElementActionCls.SafeToElementActionMeth(driver, SafeToVerify, MdataSheetTab,
                                                                         MdataSheetItem)
            if SafeToClick == "Yes":
                #driver.find_element(By.XPATH, "//button[contains(text(),'Search')]").click()
                # -------------------Property Link Text Click------------------------------
                ElementVerify = "Property [ " + SearchedEle + " ] click"
                PageName = DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PageName")
                MdataSheetTab = "test_Regression_Properties_Admi"
                MdataSheetItem = "PropertyLink"
                MdataSheetItem2 = "PropertyLinkText"
                ElementExpected = "Property Information"
                ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2,
                                                   ElementExpected,
                                                   ElementVerify,
                                                   PageName, TestResult, TestResultStatus)
                # -------------------Inspection tab------------------------------
                ElementVerify = "Inspection tab for Property [ " + SearchedEle + " ]"
                PageName = DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PageName")
                MdataSheetTab = "test_Regression_Properties_Admi"
                MdataSheetItem = "InspectionTabLink"
                MdataSheetItem2 = "InspectionTabLinkText"
                ElementExpected = "Inspection"
                ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2,
                                                   ElementExpected,
                                                   ElementVerify,
                                                   PageName, TestResult, TestResultStatus)

                # -------------------Work Orders tab------------------------------
                ElementVerify = "Work Orders tab for Property [ " + SearchedEle + " ]"
                PageName = DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PageName")
                MdataSheetTab = "test_Regression_Properties_Admi"
                MdataSheetItem = "WorkOrdersTabLink"
                MdataSheetItem2 = "WorkOrdersTabLinkText"
                ElementExpected = "Work Description"
                ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2,
                                                   ElementExpected,
                                                   ElementVerify,
                                                   PageName, TestResult, TestResultStatus)

                # -------------------Spaces and Assets tab------------------------------
                ElementVerify = "Spaces and Assets tab for Property [ " + SearchedEle + " ]"
                PageName = DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PageName")
                MdataSheetTab = "test_Regression_Properties_Admi"
                MdataSheetItem = "SpacesAndAssetsTabLink"
                MdataSheetItem2 = "SpacesAndAssetsTabLinkText"
                ElementExpected = "Spaces"
                ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2,
                                                   ElementExpected,
                                                   ElementVerify,
                                                   PageName, TestResult, TestResultStatus)
                driver.find_element(By.XPATH,
                                    DataReadMaster.GlobalData("test_Smoke_Properties_Admin", "RecordsTab")).click()
                LoaderCls.LoaderMeth(driver)
                driver.find_element(By.XPATH,
                                    DataReadMaster.GlobalData("test_Smoke_Properties_Admin", "PageLink")).click()


            # # -------------------Properties Count ------------------------------
            # ElementVerify = "Total Properties"
            # PageName = DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PageName")
            # ElementExpected = driver.find_element(By.XPATH, DataReadMaster.GlobalData("test_Regression_Properties_Admi", "TotalPropertiesCount")).text
            # MdataSheetTab = "test_Regression_Properties_Admi"
            # MdataSheetItem = "ItemsPerPageProperties"
            # ElementCountCls.ElementCountMeth(driver, MdataSheetTab, MdataSheetItem, ElementExpected, ElementVerify,
            #                                      PageName, TestResult, TestResultStatus)
            # driver.find_element(By.XPATH,
            #                     DataReadMaster.GlobalData("test_Regression_Properties_Admi", "RecordsTab")).click()
            # LoaderCls.LoaderMeth(driver)
            # driver.find_element(By.XPATH,
            #                     DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PageLink")).click()
            #
            # # -------------------Vacant Count ------------------------------
            # ElementVerify = "Total Vacant"
            # PageName = DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PageName")
            # ElementExpected = driver.find_element(By.XPATH, DataReadMaster.GlobalData("test_Regression_Properties_Admi", "TotalVacantCount")).text
            # MdataSheetTab = "test_Regression_Properties_Admi"
            # MdataSheetItem = "ItemsPerPageVacant"
            # ElementCountCls.ElementCountMeth(driver, MdataSheetTab, MdataSheetItem, ElementExpected, ElementVerify,
            #                                      PageName, TestResult, TestResultStatus)
            # driver.find_element(By.XPATH,
            #                     DataReadMaster.GlobalData("test_Regression_Properties_Admi", "RecordsTab")).click()
            # LoaderCls.LoaderMeth(driver)
            # driver.find_element(By.XPATH,
            #                     DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PageLink")).click()
            #
            # # -------------------Occupied Count ------------------------------
            # ElementVerify = "Occupied Vacant"
            # PageName = DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PageName")
            # ElementExpected = driver.find_element(By.XPATH, DataReadMaster.GlobalData("test_Regression_Properties_Admi",
            #                                                                           "TotalOccupiedCount")).text
            # MdataSheetTab = "test_Regression_Properties_Admi"
            # MdataSheetItem = "ItemsPerPageOccupied"
            # ElementCountCls.ElementCountMeth(driver, MdataSheetTab, MdataSheetItem, ElementExpected, ElementVerify,
            #                                  PageName, TestResult, TestResultStatus)
            # driver.find_element(By.XPATH,
            #                     DataReadMaster.GlobalData("test_Regression_Properties_Admi", "RecordsTab")).click()
            # LoaderCls.LoaderMeth(driver)
            # driver.find_element(By.XPATH,
            #                     DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PageLink")).click()

        except Exception as Mainerror:
            print(Mainerror)
            stringMainerror = repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror + " - "+DataReadMaster.GlobalData("test_Regression_Properties_Admi", "PageName")+" section was not able to open by automated script. Execution terminated for all other test cases")
                TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = DataReadMaster.GlobalData("test_Regression_Properties_Admi", "TestName")

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------
