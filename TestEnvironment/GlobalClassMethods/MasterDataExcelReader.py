import os
import openpyxl
import socket
class DataReadMaster:
    ROOT_DIR = None
    #print(socket.gethostname())
    if socket.gethostname()=="LAPTOP-US9BAVU1":
        ROOT_DIR = "C:/Users/Neeraj Kumar/PycharmProjects/MIMORepo/"
    if socket.gethostname() == "DESKTOP-JLLTS65":
        ROOT_DIR = "C:/Users/Neeraj/PycharmProjects/MIMO/"
    newPath = ROOT_DIR.replace(os.sep, '/')
    Path=newPath
    ExcelFileName = "MasterDataFile"
    locx = (Path + 'TestEnvironment/MasterData/' + ExcelFileName + '.xlsx')
    wbx = openpyxl.load_workbook(locx)

    @classmethod
    def GlobalData(cls,Sheetname,FieldName):
        sheetx = DataReadMaster.wbx[Sheetname]
        for ix in range(1, 200):
            if sheetx.cell(ix, 1).value == None:
                break
            else:
                if sheetx.cell(ix, 1).value == FieldName:
                    return sheetx.cell(ix, 2).value

    @classmethod
    def GlobalDataForm(cls, Sheetname, FieldName):
        sheetx = DataReadMaster.wbx[Sheetname]
        for ix in range(1, 200):
            if sheetx.cell(ix, 1).value == None:
                break
            else:
                if sheetx.cell(ix, 1).value == FieldName:
                    ColB=sheetx.cell(ix, 2).value
                    ColC=sheetx.cell(ix, 3).value
                    return ColB,ColC