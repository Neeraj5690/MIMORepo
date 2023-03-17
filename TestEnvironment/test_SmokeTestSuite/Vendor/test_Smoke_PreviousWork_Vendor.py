import datetime
import sys
if "C:/Users/Neeraj/PycharmProjects/MIMO" not in sys.path:
    sys.path.append("C:/Users/Neeraj/PycharmProjects/MIMO")
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver.common.by import By
from chrome.LatestChrome import ChromeCls
from TestEnvironment.GlobalClassMethods.MasterDataExcelReader import DataReadMaster
from TestEnvironment.GlobalElementAction.ElementAction import ElementActionCls
from TestEnvironment.GlobalElementAction.SafeToElementAction import SafeToElementActionCls, SafeToVerify
from TestEnvironment.GlobalLoader.Loader import LoaderCls

@allure.step("Entering username ")
def enter_username(username):
    driver.find_element(By.ID, "un").send_keys(username)


@allure.step("Entering password ")
def enter_password(password):
    driver.find_element(By.ID, "pw").send_keys(password)

@pytest.fixture()
def test_setup():
    global driver, TestResult, TestResultStatus, path, FundNameList, FundNameListAfterRemove, ct, Exe, D1, D2, d1, d2, DollarDate

    TestResult = []
    TestResultStatus = []
    TestFailStatus = []
    FailStatus = "Pass"
    Exe = "Yes"

    path = DataReadMaster.Path + DataReadMaster.GlobalData("test_Smoke_PreviousWork_Vendor",
                                                           "ParentDirectory") + DataReadMaster.GlobalData(
        "test_Smoke_PreviousWork_Vendor", "Directory") + DataReadMaster.GlobalData("test_Smoke_PreviousWork_Vendor", "SubDirectory")
    FundNameList = []
    FundNameListAfterRemove = []

    ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
    ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

    today = datetime.date.today()
    D1 = today.strftime("%Y-%m-%d")
    d1 = D1
    DollarDate = datetime.datetime.strptime(d1, '%Y-%m-%d')
    DollarDate = "$" + DollarDate.date().__str__() + "$"
    d1 = datetime.datetime.strptime(D1, "%Y-%m-%d")

    Exe = DataReadMaster.GlobalData("test_Smoke_PreviousWork_Vendor", "Execution")

    # --------Login to the application-----------------------
    if Exe == "Yes":
        ChromeCls.ChromeMeth()
        driver = webdriver.Chrome(executable_path=ChromeCls.NewChromePath1ChrCls)
        driver.implicitly_wait(10)
        driver.maximize_window()
        driver.get(DataReadMaster.GlobalData("GlobalData", "URLVendor"))
        enter_username(DataReadMaster.GlobalData("GlobalData", "VendorUsername"))
        enter_password(DataReadMaster.GlobalData("GlobalData", "VendorPassword"))
        driver.find_element(By.XPATH, DataReadMaster.GlobalData("GlobalData", "LoginSubmit")).click()

    yield
    if Exe == "Yes":
        class PDF(FPDF):
            def header(self):
                self.image(path + 'EmailReportContent/Logo.png', 10, 8, 33)
                self.set_font('Arial', 'B', 15)
                self.cell(73)
                self.set_text_color(0, 0, 0)
                self.cell(35, 10, ' Test Report ', 1, 1, 'B')
                self.set_font('Arial', 'I', 10)
                self.cell(150)
                self.cell(30, 10, ctReportHeader, 0, 0, 'C')
                self.ln(20)

            def footer(self):
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.set_text_color(0, 0, 0)
                self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

        pdf = PDF()
        pdf.alias_nb_pages()
        pdf.add_page()
        pdf.set_font('Times', '', 12)
        pdf.cell(0, 10, "Test Case Name:  " + DataReadMaster.GlobalData("test_Smoke_PreviousWork_Vendor", "PDFTestName"), 0, 1)
        pdf.multi_cell(0, 10,
                       "Description:  " + DataReadMaster.GlobalData("test_Smoke_PreviousWork_Vendor", "PDFDescription"), 0, 1)
        for i1 in range(len(TestResult)):
            pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(0, 0, 0)
            if (TestResultStatus[i1] == "Fail"):
                # print("Fill Red color")
                pdf.set_text_color(255, 0, 0)
                TestFailStatus.append("Fail")
            TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
            pdf.multi_cell(0, 7, str(i1 + 1) + ")  " + TestName1, 0, 1, fill=True)
            TestFailStatus.append("Pass")
        pdf.output(DataReadMaster.GlobalData("test_Smoke_PreviousWork_Vendor", "TestName") + "_" + ct + ".pdf")

        # -----------To check if any failed Test case present-------------------
        for io in range(len(TestResult)):
            if TestFailStatus[io] == "Fail":
                FailStatus = "Fail"
        # ---------------------------------------------------------------------

        # -----------To add test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        print()
        check = DataReadMaster.GlobalData("test_Smoke_PreviousWork_Vendor", "TestName")
        PdfName = DataReadMaster.GlobalData("test_Smoke_PreviousWork_Vendor", "TestName") + "_" + ct + ".pdf"
        checkcount = 0

        for i in range(1, 100):
            if sheet.cell(i, 1).value == None:
                if checkcount == 0:
                    sheet.cell(row=i, column=1).value = check
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = DataReadMaster.GlobalData("test_Smoke_PreviousWork_Vendor",
                                                                                  "TestDirectoryName")
                    sheet.cell(row=i, column=4).value = DataReadMaster.GlobalData("test_Smoke_PreviousWork_Vendor",
                                                                                  "PDFDescription")
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
                wb.save(loc)
                break
            else:
                if sheet.cell(i, 1).value == check:
                    if checkcount == 0:
                        sheet.cell(row=i, column=2).value = PdfName
                        sheet.cell(row=i, column=3).value = DataReadMaster.GlobalData("test_Smoke_PreviousWork_Vendor",
                                                                                      "TestDirectoryName")
                        sheet.cell(row=i, column=4).value = DataReadMaster.GlobalData("test_Smoke_PreviousWork_Vendor",
                                                                                      "PDFDescription")
                        sheet.cell(row=i, column=5).value = FailStatus
                        checkcount = 1
        # ----------------------------------------------------------------------------

        # ---------------------To add Test name in Execution sheet--------------------
        ExcelFileName1 = "Execution"
        loc1 = (path + 'Executiondir/' + ExcelFileName1 + '.xlsx')
        wb1 = openpyxl.load_workbook(loc1)
        sheet1 = wb1.active
        checkcount1 = 0

        for ii1 in range(1, 100):
            if sheet1.cell(ii1, 1).value == None:
                if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
                wb1.save(loc1)
                break
            else:
                if sheet1.cell(ii1, 1).value == check:
                    if checkcount1 == 0:
                        sheet1.cell(row=ii1, column=1).value = check
                        checkcount1 = 1
        # -----------------------------------------------------------------------------
        driver.quit()


@pytest.mark.smoke
def test_AllModules(test_setup):
    if Exe == "Yes":
        try:
            # ---------------------------Verify Previous Work-----------------------------
            PageName = "Previous Work"
            Ptitle1 = "Previous Work - PI - Vendor"
            driver.find_element(By.XPATH, DataReadMaster.GlobalData("test_Smoke_PreviousWork_Vendor", "PreviousWorkPage")).click()
            LoaderCls.LoaderMeth(driver)
            try:
                PageTitle1 = driver.title
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " module was not able to open"
                TestResult.append(PageName + " module opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " module was not able to open")
                TestResultStatus.append("Fail")

            # ---------------------------Verify Work Order listing click-----------------------------
            PageName = "Previous Work"
            MdataSheetTab = "test_Smoke_PreviousWork_Vendor"
            MdataSheetItem = "WorkOrdersIDClick"
            MdataSheetItem2 = "WorkOrdersIDClick1Text"
            ElementExpected = "Work Order Details"

            ElementVerify = "Work Orders Listing"
            SafeToClick = SafeToElementActionCls.SafeToElementActionMeth(driver, SafeToVerify, MdataSheetTab,
                                                                         "SafeToOpenWorkOrdersClick")
            print("SafeToClick is "+SafeToClick)
            if SafeToClick == "Yes":
                # -------------------Work OrderID Click------------------------------
                Element = "Work Order ID"
                Click_Element = "WorkOrdersID_Ele"
                if Click_Element == "Skip":
                    ElementVerify = Element + " click for " + ElementVerify
                else:
                    ElementID = driver.find_element(By.XPATH,
                                                    DataReadMaster.GlobalData(MdataSheetTab, Click_Element)).text
                    ElementVerify = Element + " [ " + ElementID + " ] click for " + ElementVerify

                ElementActionCls.ElementActionMeth(driver,MdataSheetTab, MdataSheetItem, MdataSheetItem2,
                                                   ElementExpected,
                                                   ElementVerify,
                                                   PageName, TestResult, TestResultStatus)
                # ---------------------------Work Order Details tabs-----------------------------
                # -------------------Summary Tab------------------------------
                ElementVerify = "Summary Tab"
                PageName = "Previous Work"
                MdataSheetTab = "test_Smoke_PreviousWork_Vendor"
                MdataSheetItem = "SummaryTabClick"
                MdataSheetItem2 = "SummaryTabClickText"
                ElementExpected = "Work Order Details"
                ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2,
                                                   ElementExpected,
                                                   ElementVerify,
                                                   PageName, TestResult, TestResultStatus)

                # -------------------Vendor Summary Tab------------------------------
                ElementVerify = "Vendor Summary Tab"
                PageName = "Previous Work"
                MdataSheetTab = "test_Smoke_PreviousWork_Vendor"
                MdataSheetItem = "VendorSummaryTabClick"
                MdataSheetItem2 = "VendorSummaryTabClickText"
                ElementExpected = "Work Order Details"
                ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2,
                                                   ElementExpected,
                                                   ElementVerify,
                                                   PageName, TestResult, TestResultStatus)

                # -------------------Comment History Tab------------------------------
                ElementVerify = "Comment History Tab"
                PageName = "Previous Work"
                MdataSheetTab = "test_Smoke_PreviousWork_Vendor"
                MdataSheetItem = "CommentHistoryTabClick"
                MdataSheetItem2 = "CommentHistoryTabClickText"
                ElementExpected = "CREATES A COMMENT TO BE ADDED TO THE AUDIT TABLE."
                ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2,
                                                   ElementExpected,
                                                   ElementVerify,
                                                   PageName, TestResult, TestResultStatus)

                # -------------------Related Action Tab------------------------------
                ElementVerify = "Related Action Tab"
                PageName = "Previous Work"
                MdataSheetTab = "test_Smoke_PreviousWork_Vendor"
                MdataSheetItem = "RelatedActionTabClick"
                MdataSheetItem2 = "RelatedActionTabClickText"
                ElementExpected = "Create Comment"
                ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2,
                                                   ElementExpected,
                                                   ElementVerify,
                                                   PageName, TestResult, TestResultStatus)

            else:
                print("No Data available **************** for " + ElementVerify)
            driver.find_element(By.XPATH, DataReadMaster.GlobalData("test_Smoke_PreviousWork_Vendor", "PreviousWorkPage")).click()

        except Exception as Mainerror:
            print(Mainerror)
            stringMainerror = repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror)
                TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = DataReadMaster.GlobalData("test_Smoke_PreviousWork_Vendor", "TestName")

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------