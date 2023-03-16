import os
import smtplib
import ssl
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
import pandas as pd
import ast
import sys
if "C:/Users/Neeraj/PycharmProjects/MIMO" not in sys.path:
    sys.path.append("C:/Users/Neeraj/PycharmProjects/MIMO")
from TestEnvironment.GlobalClassMethods.MasterDataExcelReader import DataReadMaster

def test_ReportSendSmokeAll():
    DirectoryName = []
    PDFName1=[]
    TestName=[]
    TestName1=[]
    TestDescription = []
    TestStatus = []
    SendStatus = []
    AttachmentAdded= []

    PDFpath = DataReadMaster.Path + DataReadMaster.GlobalData("test_Smoke_Home_Superintendent", "ParentDirectory") + DataReadMaster.GlobalData("test_Smoke_Home_Superintendent", "Directory")+ DataReadMaster.GlobalData("test_Smoke_Home_Superintendent", "SubDirectory")

    #-------------------To read content to send in e-Mail--------------------
    ExcelFileName = "FileName"
    loc = (PDFpath+'PDFFileNameData/' + ExcelFileName + '.xlsx')
    wb=openpyxl.load_workbook(loc)
    sheet = wb.active
    for i in range(1, 100):
        if sheet.cell(i, 1).value == None:
            break
        else:
            PDFName1.append(sheet.cell(i, 2).value)
            DirectoryName.append(sheet.cell(i, 3).value)
            TestName.append(sheet.cell(i, 1).value+".pdf")
            TestName1.append(sheet.cell(i, 1).value)
            TestDescription.append(sheet.cell(i, 4).value)
            TestStatus.append(sheet.cell(i, 5).value)
            SendStatus.append(sheet.cell(i, 6).value)

        B = ""
        for io in range(len(TestName)):
            try:
                B = B + "<br /><br />"+str(io+1)+") " + "".join(TestName1[io])+" => "+"".join(TestDescription[io])+" => "+"".join(TestStatus[io])
            except Exception:
                print("No attachment details to add in email description")

    html = '''
        <html>
            <body>
                <p>Hi Team <br />'''+DataReadMaster.GlobalData('GlobalData', 'EmailIntro_Smoke_Superintendent')+'''<br />Below 
                test scenarios are covered</p> <p></p> <p>'''+B+'''</p 
                <p></p>
                <img src='cid:myimageid' width="500" align="center">
                <p>Please find attached PDFs for detailed information on test scenarios results<br /><br /></p>
                <p>Many Thanks <br/>Neeraj</p>
            </body>
        </html>
        '''

    def attach_file_to_email(msg, attach,filename, extra_headers=None):
        with open(attach, "rb") as f:
            file_attachment = MIMEApplication(f.read())
        file_attachment.add_header(
            "Content-Disposition",
            f"attachment; filename= {filename}",
        )
        if extra_headers is not None:
            for name, value in extra_headers.items():
                file_attachment.add_header(name, value)
        msg.attach(file_attachment)

    email_from = 'Test Automation Team '
    y = DataReadMaster.GlobalData("GlobalData", "EmailTo")
    email_to = ast.literal_eval(y)
    SenderEmail=DataReadMaster.GlobalData("GlobalData", "EmailFrom")
    date_str = pd.Timestamp.today().strftime('%m-%d-%Y')
    msg = MIMEMultipart()
    msg['Subject']=DataReadMaster.GlobalData("GlobalData", "EmailSubject_Smoke_Superintendent")+' -Test Automation Report- '+date_str
    msg['From'] = email_from
    msg['To'] = ','.join(email_to)
    msg.attach(MIMEText(html, "html"))

    #-----------------------------------------------------------------------
    try:
        attach_file_to_email(msg, PDFpath+'TestPieResult.png',"Pie.png",
                         {'Content-ID': '<myimageid>'})
    except Exception:
        print("No Pie File to attach")
    # ------------------To add attachments in the report email--------------
    i=0
    for file in PDFName1:
        print()
        try:
            #print(file)
            if SendStatus[i]=="Send Only when Fail=Yes" and  TestStatus[i]=="Fail":
                    attach_file_to_email(msg,PDFpath+PDFName1[i],TestName[i])
                    print(TestName[i])
                    AttachmentAdded.append("Yes")
            if SendStatus[i] == "Send Only when Fail=No":
                    print("Inside Send Only when Fail=No")
                    attach_file_to_email(msg, PDFpath+PDFName1[i],TestName[i])
                    print(TestName[i])
                    AttachmentAdded.append("Yes")
        except Exception as e1:
            print("No Attachment found to Add")
            print(e1)
        i = i + 1
    #-----------------------------------------------------------------------

    # ------------------------To attach all in e-Mail-----------------------
    email_string = msg.as_string()
    context = ssl.create_default_context()
    # -----------------------------------------------------------------------

    # ----------------------------SMTP setup--------------------------------
    server=smtplib.SMTP_SSL('smtp.gmail.com',465)
    RandmStr=DataReadMaster.GlobalData("GlobalData", "GoogleAppCode")
    server.login(SenderEmail,RandmStr)
    #-----------------------------------------------------------------------

    #---------------------------------Sending email-------------------------
    for io1 in range(len(AttachmentAdded)):
        if AttachmentAdded[io1] == "Yes":
            print("Inside Attachment Added=Yes ")
            server.sendmail(email_from, email_to, email_string)
            print("Test Report sent")
            break
    #-----------------------------------------------------------------------

    #-----------------To delete pdf and report files----------------------------
    ii=0
    for ii in range(0,len(PDFName1)):
        print()
        try:
            os.remove(PDFpath+PDFName1[ii])
        except Exception as eer:
            print(eer)
            print("No Attachment found to delete")
    try:
        os.remove(PDFpath+'TestPieResult.png')
    except Exception:
        print("No Attachment found to delete")
    #-----------------------------------------------------------------------
    server.quit()
