import datetime
import sys
if "C:/Users/Neeraj/PycharmProjects/MIMO" not in sys.path:
    sys.path.append("C:/Users/Neeraj/PycharmProjects/MIMO")
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver.common.by import By
from chrome.LatestChrome import ChromeCls
from TestEnvironment.GlobalClassMethods.MasterDataExcelReader import DataReadMaster
from TestEnvironment.GlobalElementAction.ElementAction import ElementActionCls
from TestEnvironment.GlobalElementAction.SafeToElementAction import SafeToElementActionCls, SafeToVerify
from TestEnvironment.GlobalElementPresent.ElementPresent import ElementPresentCls
from TestEnvironment.GlobalLoader.Loader import LoaderCls


@allure.step("Entering username ")
def enter_username(username):
    driver.find_element(By.ID, "un").send_keys(username)


@allure.step("Entering password ")
def enter_password(password):
    driver.find_element(By.ID, "pw").send_keys(password)


@pytest.fixture()
def test_setup():
    sys.path.append("/chrome")
    print(ChromeCls.NewChromePathChrCls)

    global driver, TestResult, TestResultStatus, path, FundNameList, FundNameListAfterRemove, ct, Exe, D1, D2, d1, d2, DollarDate

    TestResult = []
    TestResultStatus = []
    TestFailStatus = []
    FailStatus = "Pass"
    Exe = "Yes"

    path = DataReadMaster.Path + DataReadMaster.GlobalData("test_Smoke_Home_Admin",
                                                           "ParentDirectory") + DataReadMaster.GlobalData(
        "test_Smoke_Home_Admin", "Directory") + DataReadMaster.GlobalData("test_Smoke_Home_Admin", "SubDirectory")
    FundNameList = []
    FundNameListAfterRemove = []

    ct = datetime.datetime.now().strftime("%d_%B_%Y_%I_%M%p")
    ctReportHeader = datetime.datetime.now().strftime("%d %B %Y %I %M%p")

    today = datetime.date.today()
    D1 = today.strftime("%Y-%m-%d")
    d1 = D1
    DollarDate = datetime.datetime.strptime(d1, '%Y-%m-%d')
    DollarDate = "$" + DollarDate.date().__str__() + "$"
    d1 = datetime.datetime.strptime(D1, "%Y-%m-%d")

    Exe = DataReadMaster.GlobalData("test_Smoke_Home_Admin", "Execution")

    # --------Login to the application-----------------------
    if Exe == "Yes":
        ChromeCls.ChromeMeth()
        driver = webdriver.Chrome(executable_path=ChromeCls.NewChromePath1ChrCls)
        driver.implicitly_wait(10)
        driver.maximize_window()
        driver.get(DataReadMaster.GlobalData("GlobalData", "URL1"))
        enter_username(DataReadMaster.GlobalData("GlobalData", "AdminUsername"))
        enter_password(DataReadMaster.GlobalData("GlobalData", "AdminPassword"))
        driver.find_element(By.XPATH, DataReadMaster.GlobalData("GlobalData", "LoginSubmit")).click()

    yield
    if Exe == "Yes":
        class PDF(FPDF):
            def header(self):
                self.image(path + 'EmailReportContent/Logo.png', 10, 8, 33)
                self.set_font('Arial', 'B', 15)
                self.cell(73)
                self.set_text_color(0, 0, 0)
                self.cell(35, 10, ' Test Report ', 1, 1, 'B')
                self.set_font('Arial', 'I', 10)
                self.cell(150)
                self.cell(30, 10, ctReportHeader, 0, 0, 'C')
                self.ln(20)

            def footer(self):
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.set_text_color(0, 0, 0)
                self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

        pdf = PDF()
        pdf.alias_nb_pages()
        pdf.add_page()
        pdf.set_font('Times', '', 12)
        pdf.cell(0, 10, "Test Case Name:  " + DataReadMaster.GlobalData("test_Smoke_Home_Admin", "PDFTestName"), 0, 1)
        pdf.multi_cell(0, 10,
                       "Description:  " + DataReadMaster.GlobalData("test_Smoke_Home_Admin", "PDFDescription"), 0, 1)
        for i1 in range(len(TestResult)):
            pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(0, 0, 0)
            if (TestResultStatus[i1] == "Fail"):
                # print("Fill Red color")
                pdf.set_text_color(255, 0, 0)
                TestFailStatus.append("Fail")
            TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
            pdf.multi_cell(0, 7, str(i1 + 1) + ")  " + TestName1, 0, 1, fill=True)
            TestFailStatus.append("Pass")
        pdf.output(DataReadMaster.GlobalData("test_Smoke_Home_Admin", "TestName") + "_" + ct + ".pdf")

        # -----------To check if any failed Test case present-------------------
        for io in range(len(TestResult)):
            if TestFailStatus[io] == "Fail":
                FailStatus = "Fail"
        # ---------------------------------------------------------------------

        # -----------To add test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        print()
        check = DataReadMaster.GlobalData("test_Smoke_Home_Admin", "TestName")
        PdfName = DataReadMaster.GlobalData("test_Smoke_Home_Admin", "TestName") + "_" + ct + ".pdf"
        checkcount = 0

        for i in range(1, 100):
            if sheet.cell(i, 1).value == None:
                if checkcount == 0:
                    sheet.cell(row=i, column=1).value = check
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = DataReadMaster.GlobalData("test_Smoke_Home_Admin",
                                                                                  "TestDirectoryName")
                    sheet.cell(row=i, column=4).value = DataReadMaster.GlobalData("test_Smoke_Home_Admin",
                                                                                  "PDFDescription")
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
                wb.save(loc)
                break
            else:
                if sheet.cell(i, 1).value == check:
                    if checkcount == 0:
                        sheet.cell(row=i, column=2).value = PdfName
                        sheet.cell(row=i, column=3).value = DataReadMaster.GlobalData("test_Smoke_Home_Admin",
                                                                                      "TestDirectoryName")
                        sheet.cell(row=i, column=4).value = DataReadMaster.GlobalData("test_Smoke_Home_Admin",
                                                                                      "PDFDescription")
                        sheet.cell(row=i, column=5).value = FailStatus
                        checkcount = 1
        # ----------------------------------------------------------------------------

        # ---------------------To add Test name in Execution sheet--------------------
        ExcelFileName1 = "Execution"
        loc1 = (path + 'Executiondir/' + ExcelFileName1 + '.xlsx')
        wb1 = openpyxl.load_workbook(loc1)
        sheet1 = wb1.active
        checkcount1 = 0

        for ii1 in range(1, 100):
            if sheet1.cell(ii1, 1).value == None:
                if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
                wb1.save(loc1)
                break
            else:
                if sheet1.cell(ii1, 1).value == check:
                    if checkcount1 == 0:
                        sheet1.cell(row=ii1, column=1).value = check
                        checkcount1 = 1
        # -----------------------------------------------------------------------------
        driver.quit()


@pytest.mark.smoke
def test_AllModules(test_setup):
    if Exe == "Yes":
        try:
            # ---------------------------Verify Home page-----------------------------
            PageName = "Home"
            Ptitle1 = "Home - PI - Administrator"
            print(driver.title)
            time.sleep(1)
            LoaderCls.LoaderMeth(driver)
            try:
                PageTitle1 = driver.title
                assert PageTitle1 in Ptitle1, PageName + " module was not able to open"
                TestResult.append(PageName + " module opened successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " module was not able to open")
                TestResultStatus.append("Fail")

            # --------------------Checking Inspections Graphs at Home--------------
            # -------------------Inspections New Graph------------------------------
            driver.find_element(By.XPATH,
                                DataReadMaster.GlobalData("test_Smoke_Home_Admin", "HomePage")).click()
            ElementVerify = "Inspections New Graph"
            PageName = "Home"
            ElementExpected = "New"
            MdataSheetTab = "test_Smoke_Home_Admin"
            MdataSheetItem = "InspectionsGraphNew"
            ElementPresentCls.ElementPresentMeth(driver, MdataSheetTab, MdataSheetItem, ElementExpected, ElementVerify,
                                                 PageName, TestResult, TestResultStatus)

            # -------------------Inspections Completed Graph------------------------------
            ElementVerify = "Inspections Completed Graph"
            PageName = "Home"
            ElementExpected = "Completed"
            MdataSheetTab = "test_Smoke_Home_Admin"
            MdataSheetItem = "InspectionsGraphCompleted"
            ElementPresentCls.ElementPresentMeth(driver, MdataSheetTab, MdataSheetItem, ElementExpected, ElementVerify,
                                                 PageName, TestResult, TestResultStatus)

            # -------------------Work Orders New Graph------------------------------
            ElementVerify = "Work Orders New Graph"
            PageName = "Home"
            ElementExpected = "New"
            MdataSheetTab = "test_Smoke_Home_Admin"
            MdataSheetItem = "WorkOrdersGraphNew"
            ElementPresentCls.ElementPresentMeth(driver, MdataSheetTab, MdataSheetItem, ElementExpected, ElementVerify,
                                                 PageName, TestResult, TestResultStatus)

            # -------------------Work Orders Completed Graph------------------------------
            ElementVerify = "Work Orders Completed Graph"
            PageName = "Home"
            ElementExpected = "Completed"
            MdataSheetTab = "test_Smoke_Home_Admin"
            MdataSheetItem = "WorkOrdersGraphCompleted"
            ElementPresentCls.ElementPresentMeth(driver, MdataSheetTab, MdataSheetItem, ElementExpected, ElementVerify,
                                                 PageName, TestResult, TestResultStatus)

            # --------------------Some Tab BUTTONS------------------------
            # -------------------Incomplete Inspections Button------------------------------
            ElementVerify = "Incomplete Inspections Button"
            PageName = "Home"
            MdataSheetTab = "test_Smoke_Home_Admin"
            MdataSheetItem = "IncompleteInspectionsButton"
            MdataSheetItem2 = "IncompleteInspectionsButtonText"
            ElementExpected = "Inspections"
            ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2, ElementExpected,
                                               ElementVerify,
                                               PageName, TestResult, TestResultStatus)
            SafeToClick = SafeToElementActionCls.SafeToElementActionMeth(driver, SafeToVerify, "test_Smoke_Home_Admin",
                                                                         "SafeToIncompleteInspectionsClick")
            Count = driver.find_element(By.XPATH,
                                        DataReadMaster.GlobalData("test_Smoke_Home_Admin", MdataSheetItem)).text
            start = '('
            end = ')'
            Count = Count[Count.find(start) + len(start):Count.rfind(end)]
            print("Count is " + Count)
            if SafeToClick == "Yes":
                # -------------------Incomplete Inspections Count------------------------------

                try:
                    print("11")
                    FooterCount = driver.find_element(By.XPATH,
                                                      "//aspan[contains(text(),'Properties With Open Work')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div/span[2]/strong").text
                except:
                    try:
                        print("22")
                        FooterCount = driver.find_element(By.XPATH,
                                                          "//aspan[contains(text(),'Properties With Open Work')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div/span[3]").text
                        start = 'of '
                        end = ''
                        FooterCount = FooterCount[FooterCount.find(start) + len(start):FooterCount.rfind(end)]
                    except:
                        try:
                            print("333")
                            FooterCount = driver.find_element(By.XPATH,
                                                              "//aspan[contains(text(),'Properties With Open Work')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[1]/table/thead/tr/th[2]/div[contains(text(),'Inspections')]/parent::th/parent::tr/parent::thead/parent::table/tbody/tr[1]/td[1]").text
                            if "No" or "no" in FooterCount:
                                print("No found, so count is 0")
                                FooterCount = "0"

                        except:
                            print("444")
                            FooterCount = driver.find_elements(By.XPATH,
                                                              "//span[contains(text(),'Properties With Open Work')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[1]/table/tbody/tr")
                            FooterCount=str(len(FooterCount))

                # -------------------Home Property Click------------------------------
                ElementVerify = "Property link text click for " + ElementVerify
                PageName = "Home"
                MdataSheetTab = "test_Smoke_Home_Admin"
                MdataSheetItem = "HomePropertyClick"
                MdataSheetItem2 = "HomePropertyClickText"
                ElementExpected = "Property Information"
                ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2,
                                                   ElementExpected,
                                                   ElementVerify,
                                                   PageName, TestResult, TestResultStatus)
                driver.find_element(By.XPATH, "//a/div[contains(text(),'Home')]").click()
            else:
                print("No Data available **************** for " + ElementVerify)
                FooterCount="0"
            #---------------Matching Count for Incomplete Inspections ---------------
            ElementVerify = "Incomplete Inspections Count"
            print("FooterCount is " + FooterCount)
            ElementFound = FooterCount
            if Count == ElementFound:
                TestResult.append(
                    ElementVerify + " (" + Count + ") " + " at " + PageName + " was matching with number of records" + " (" + ElementFound + ")")
                TestResultStatus.append("Pass")
            else:
                TestResult.append(
                    ElementVerify + " (" + Count + ") " + " at " + PageName + "was not matching with number of "
                                                                                        "records" + " (" +
                    ElementFound + ")")
                TestResultStatus.append("Fail")

            # -------------------Pending Inspections Button------------------------------
            ElementVerify = "Pending Inspections Button"
            PageName = "Home"
            MdataSheetTab = "test_Smoke_Home_Admin"
            MdataSheetItem = "PendingInspectionsButton"
            MdataSheetItem2 = "PendingInspectionsButtonText"
            ElementExpected = "Inspections"
            ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2, ElementExpected,
                                               ElementVerify,
                                               PageName, TestResult, TestResultStatus)
            SafeToClick = SafeToElementActionCls.SafeToElementActionMeth(driver, SafeToVerify, "test_Smoke_Home_Admin",
                                                                         "SafeToPendingInspectionsClick")
            Count = driver.find_element(By.XPATH,
                                        DataReadMaster.GlobalData("test_Smoke_Home_Admin", MdataSheetItem)).text
            start = '('
            end = ')'
            Count = Count[Count.find(start) + len(start):Count.rfind(end)]
            print("Count is " + Count)
            if SafeToClick == "Yes":
                # -------------------Pending Inspections Count------------------------------

                try:
                    print("11")
                    FooterCount = driver.find_element(By.XPATH,
                                                      "//span[contains(text(),'Properties With Open Work')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div/span[2]/strong").text
                except:
                    try:
                        print("22")
                        FooterCount = driver.find_element(By.XPATH,
                                                          "//span[contains(text(),'Properties With Open Work')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div/span[3]").text
                        start = 'of '
                        end = ''
                        FooterCount = FooterCount[FooterCount.find(start) + len(start):FooterCount.rfind(end)]
                    except:
                        try:
                            print("333")
                            FooterCount = driver.find_element(By.XPATH,
                                                              "//span[contains(text(),'Properties With Open Work')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[1]/table/thead/tr/th[2]/div[contains(text(),'Inspections')]/parent::th/parent::tr/parent::thead/parent::table/tbody/tr[1]/td[1]").text
                            if "No" or "no" in FooterCount:
                                print("No found, so count is 0")
                                FooterCount = "0"
                        except:
                            print("444")
                            FooterCount = driver.find_elements(By.XPATH,
                                                               "//span[contains(text(),'Properties With Open Work')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[1]/table/tbody/tr")
                            FooterCount = str(len(FooterCount))

                # -------------------Home Property Click------------------------------
                ElementVerify = "Property link text click for " + ElementVerify
                PageName = "Home"
                MdataSheetTab = "test_Smoke_Home_Admin"
                MdataSheetItem = "HomePropertyClick"
                MdataSheetItem2 = "HomePropertyClickText"
                ElementExpected = "Property Information"
                ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2,
                                                   ElementExpected,
                                                   ElementVerify,
                                                   PageName, TestResult, TestResultStatus)
                driver.find_element(By.XPATH, "//a/div[contains(text(),'Home')]").click()
            else:
                print("No Data available **************** for " + ElementVerify)
                FooterCount="0"
                # ---------------Matching Count for Pending Inspections ---------------
            ElementVerify = "Pending Inspections Count"
            print("FooterCount is " + FooterCount)
            ElementFound = FooterCount
            if Count == ElementFound:
                TestResult.append(
                    ElementVerify + " (" + Count + ") " + " at " + PageName + " was matching with number of records" + " (" + ElementFound + ")")
                TestResultStatus.append("Pass")
            else:
                TestResult.append(
                    ElementVerify + " (" + Count + ") " + " at " + PageName + "was not matching with number of "
                                                                                        "records" + " (" +
                    ElementFound + ")")
                TestResultStatus.append("Fail")
            # -------------------Open Work Orders Button------------------------------
            ElementVerify = "Open Work Orders Button"
            PageName = "Home"
            MdataSheetTab = "test_Smoke_Home_Admin"
            MdataSheetItem = "OpenWorkOrdersButton"
            MdataSheetItem2 = "OpenWorkOrdersButtonText"
            ElementExpected = "Work Orders"
            ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2, ElementExpected,
                                               ElementVerify,
                                               PageName, TestResult, TestResultStatus)
            SafeToClick = SafeToElementActionCls.SafeToElementActionMeth(driver, SafeToVerify, "test_Smoke_Home_Admin",
                                                                         "SafeToOpenWorkOrdersClick")
            Count = driver.find_element(By.XPATH,
                                        DataReadMaster.GlobalData("test_Smoke_Home_Admin", MdataSheetItem)).text
            start = '('
            end = ')'
            Count = Count[Count.find(start) + len(start):Count.rfind(end)]
            print("Count is " + Count)
            if SafeToClick == "Yes":
                # -------------------Open Work Orders Count------------------------------

                try:
                    print("11")
                    FooterCount = driver.find_element(By.XPATH,
                                                      "//span[contains(text(),'Properties With Open Work')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div/span[2]/strong").text
                except:
                    try:
                        print("22")
                        FooterCount = driver.find_element(By.XPATH,
                                                          "//span[contains(text(),'Properties With Open Work')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div/span[3]").text
                        start = 'of '
                        end = ''
                        FooterCount = FooterCount[FooterCount.find(start) + len(start):FooterCount.rfind(end)]
                    except:
                        try:
                            print("333")
                            FooterCount = driver.find_element(By.XPATH,
                                                              "//span[contains(text(),'Properties With Open Work')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[1]/table/thead/tr/th[2]/div[contains(text(),'Inspections')]/parent::th/parent::tr/parent::thead/parent::table/tbody/tr[1]/td[1]").text
                            if "No" or "no" in FooterCount:
                                print("No found, so count is 0")
                                FooterCount = "0"
                        except:
                            print("444")
                            FooterCount = driver.find_elements(By.XPATH,
                                                               "//span[contains(text(),'Properties With Open Work')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[1]/table/tbody/tr")
                            FooterCount = str(len(FooterCount))
                # -------------------Home Property Click------------------------------
                ElementVerify = "Property link text click for " + ElementVerify
                PageName = "Home"
                MdataSheetTab = "test_Smoke_Home_Admin"
                MdataSheetItem = "HomePropertyClick"
                MdataSheetItem2 = "HomePropertyClickText"
                ElementExpected = "Property Information"
                ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2,
                                                   ElementExpected,
                                                   ElementVerify,
                                                   PageName, TestResult, TestResultStatus)
                driver.find_element(By.XPATH, "//a/div[contains(text(),'Home')]").click()
            else:
                print("No Data available **************** for " + ElementVerify)
                FooterCount="0"
            # ---------------Matching Count for Open Work Orders ---------------
            ElementVerify = "Open Work Orders Count"
            print("FooterCount is " + FooterCount)
            ElementFound = FooterCount
            if Count == ElementFound:
                TestResult.append(
                    ElementVerify + " (" + Count + ") " + " at " + PageName + " was matching with number of records" + " (" + ElementFound + ")")
                TestResultStatus.append("Pass")
            else:
                TestResult.append(
                    ElementVerify + " (" + Count + ") " + " at " + PageName + "was not matching with number of "
                                                                                        "records" + " (" +
                    ElementFound + ")")
                TestResultStatus.append("Fail")

            # --------------------Inspections Requiring Action BUTTONS------------------------
            # -------------------Work Order Required Button------------------------------
            ElementVerify = "Work Order Required Button"
            PageName = "Home"
            MdataSheetTab = "test_Smoke_Home_Admin"
            MdataSheetItem = "WorkOrderRequiredButton"
            MdataSheetItem2 = "WorkOrderRequiredButtonText"
            ElementExpected = "Inspection Type"
            ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2, ElementExpected,
                                               ElementVerify,
                                               PageName, TestResult, TestResultStatus)
            SafeToClick = SafeToElementActionCls.SafeToElementActionMeth(driver, SafeToVerify, "test_Smoke_Home_Admin",
                                                                         "SafeToWorkOrderRequiredClick")
            Count = driver.find_element(By.XPATH,
                                        DataReadMaster.GlobalData("test_Smoke_Home_Admin", MdataSheetItem)).text
            start = '('
            end = ')'
            Count = Count[Count.find(start) + len(start):Count.rfind(end)]
            print("Count is " + Count)
            if SafeToClick == "Yes":
                # -------------------Work Order Required Count------------------------------
                try:
                    print("11")
                    FooterCount = driver.find_element(By.XPATH,
                                                      "//span[contains(text(),'Inspections Requiring Action')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div/span[2]/strong").text
                except:
                    try:
                        print("22")
                        FooterCount = driver.find_element(By.XPATH,
                                                          "//span[contains(text(),'Inspections Requiring Action')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div/span[3]").text
                        start = 'of '
                        end = ''
                        FooterCount = FooterCount[FooterCount.find(start) + len(start):FooterCount.rfind(end)]
                    except:
                        try:
                            print("333")
                            FooterCount = driver.find_element(By.XPATH,
                                                              "//span[contains(text(),'Inspections Requiring Action')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[1]/table/thead/tr/th[2]/div[contains(text(),'Inspections')]/parent::th/parent::tr/parent::thead/parent::table/tbody/tr[1]/td[1]").text
                            if "No" or "no" in FooterCount:
                                print("No found, so count is 0")
                                FooterCount = "0"
                        except:
                            print("444")
                            FooterCount = driver.find_elements(By.XPATH,
                                                               "//span[contains(text(),'Inspections Requiring Action')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[1]/table/tbody/tr")
                            FooterCount = str(len(FooterCount))
                # -------------------Home Inspection Click------------------------------
                ElementVerify = "Inspection link text click for " + ElementVerify
                PageName = "Home"
                MdataSheetTab = "test_Smoke_Home_Admin"
                MdataSheetItem = "HomeInspectionClick"
                MdataSheetItem2 = "HomeInspectionClickText"
                ElementExpected = "Inspection Date"
                ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2,
                                                   ElementExpected,
                                                   ElementVerify,
                                                   PageName, TestResult, TestResultStatus)
                driver.find_element(By.XPATH, "//a/div[contains(text(),'Home')]").click()
            else:
                print("No Data available **************** for " + ElementVerify)
                FooterCount="0"
            # ---------------Matching Count for Work Order Required ---------------
            ElementVerify = "Work Order Required Count"
            print("FooterCount is " + FooterCount)
            ElementFound = FooterCount
            if Count == ElementFound:
                TestResult.append(
                    ElementVerify + " (" + Count + ") " + " at " + PageName + " was matching with number of records" + " (" + ElementFound + ")")
                TestResultStatus.append("Pass")
            else:
                TestResult.append(
                    ElementVerify + " (" + Count + ") " + " at " + PageName + "was not matching with number of "
                                                                                        "records" + " (" +
                    ElementFound + ")")
                TestResultStatus.append("Fail")

            # -------------------Sign off Required Button------------------------------
            ElementVerify = "Sign Off Required Button"
            PageName = "Home"
            MdataSheetTab = "test_Smoke_Home_Admin"
            MdataSheetItem = "SignOffRequiredButton"
            MdataSheetItem2 = "SignOffRequiredButtonText"
            ElementExpected = "Inspection Type"
            ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2, ElementExpected,
                                               ElementVerify,
                                               PageName, TestResult, TestResultStatus)
            SafeToClick = SafeToElementActionCls.SafeToElementActionMeth(driver, SafeToVerify, "test_Smoke_Home_Admin",
                                                                         "SafeToSignOffRequiredClick")
            Count = driver.find_element(By.XPATH,
                                        DataReadMaster.GlobalData("test_Smoke_Home_Admin", MdataSheetItem)).text
            start = '('
            end = ')'
            Count = Count[Count.find(start) + len(start):Count.rfind(end)]
            print("Count is " + Count)
            if SafeToClick == "Yes":
                # -------------------Sign Off Required Count------------------------------

                try:
                    print("11")
                    FooterCount = driver.find_element(By.XPATH,
                                                      "//span[contains(text(),'Inspections Requiring Action')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div/span[2]/strong").text
                except:
                    try:
                        print("22")
                        FooterCount = driver.find_element(By.XPATH,
                                                          "//span[contains(text(),'Inspections Requiring Action')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div/span[3]").text
                        start = 'of '
                        end = ''
                        FooterCount = FooterCount[FooterCount.find(start) + len(start):FooterCount.rfind(end)]
                    except:
                        try:
                            print("333")
                            FooterCount = driver.find_element(By.XPATH,
                                                              "//span[contains(text(),'Inspections Requiring Action')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[1]/table/thead/tr/th[2]/div[contains(text(),'Inspections')]/parent::th/parent::tr/parent::thead/parent::table/tbody/tr[1]/td[1]").text
                            if "No" or "no" in FooterCount:
                                print("No found, so count is 0")
                                FooterCount = "0"
                        except:
                            print("444")
                            FooterCount = driver.find_elements(By.XPATH,
                                                               "//span[contains(text(),'Inspections Requiring Action')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[1]/table/tbody/tr")
                            FooterCount = str(len(FooterCount))
                # -------------------Home Inspection Click------------------------------
                ElementVerify = "Inspection link text click for " + ElementVerify
                PageName = "Home"
                MdataSheetTab = "test_Smoke_Home_Admin"
                MdataSheetItem = "HomeInspectionClick"
                MdataSheetItem2 = "HomeInspectionClickText"
                ElementExpected = "Inspection Date"
                ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2,
                                                   ElementExpected,
                                                   ElementVerify,
                                                   PageName, TestResult, TestResultStatus)
                driver.find_element(By.XPATH, "//a/div[contains(text(),'Home')]").click()
            else:
                print("No Data available **************** for " + ElementVerify)
                FooterCount="0"
            # ---------------Matching Count for Sign Off Required ---------------
            ElementVerify = "Sign Off Required Count"
            print("FooterCount is " + FooterCount)
            ElementFound = FooterCount
            if Count == ElementFound:
                TestResult.append(
                    ElementVerify + " (" + Count + ") " + " at " + PageName + " was matching with number of records" + " (" + ElementFound + ")")
                TestResultStatus.append("Pass")
            else:
                TestResult.append(
                    ElementVerify + " (" + Count + ") " + " at " + PageName + "was not matching with number of "
                                                                                        "records" + " (" +
                    ElementFound + ")")
                TestResultStatus.append("Fail")

            # -------------------Finalization Required Button------------------------------
            ElementVerify = "Finalization Required Button"
            PageName = "Home"
            MdataSheetTab = "test_Smoke_Home_Admin"
            MdataSheetItem = "FinalizationRequiredButton"
            MdataSheetItem2 = "FinalizationRequiredButtonText"
            ElementExpected = "Inspection Type"
            ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2, ElementExpected,
                                               ElementVerify,
                                               PageName, TestResult, TestResultStatus)
            SafeToClick = SafeToElementActionCls.SafeToElementActionMeth(driver, SafeToVerify, "test_Smoke_Home_Admin",
                                                                         "SafeToFinalizationRequiredClick")
            Count = driver.find_element(By.XPATH,
                                        DataReadMaster.GlobalData("test_Smoke_Home_Admin", MdataSheetItem)).text
            start = '('
            end = ')'
            Count = Count[Count.find(start) + len(start):Count.rfind(end)]
            print("Count is " + Count)
            if SafeToClick=="Yes":
                # -------------------Finalization Required Count------------------------------

                try:
                    print("11")
                    FooterCount = driver.find_element(By.XPATH,
                                                      "//span[contains(text(),'Inspections Requiring Action')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div/span[2]/strong").text
                except:
                    try:
                        print("22")
                        FooterCount = driver.find_element(By.XPATH,
                                                          "//span[contains(text(),'Inspections Requiring Action')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div/span[3]").text
                        start = 'of '
                        end = ''
                        FooterCount = FooterCount[FooterCount.find(start) + len(start):FooterCount.rfind(end)]
                    except:
                        try:
                            print("333")
                            FooterCount = driver.find_element(By.XPATH,
                                                              "//span[contains(text(),'Inspections Requiring Action')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[1]/table/thead/tr/th[2]/div[contains(text(),'Inspections')]/parent::th/parent::tr/parent::thead/parent::table/tbody/tr[1]/td[1]").text
                            if "No" or "no" in FooterCount:
                                print("No found, so count is 0")
                                FooterCount = "0"
                        except:
                            print("444")
                            FooterCount = driver.find_elements(By.XPATH,
                                                               "//span[contains(text(),'Inspections Requiring Action')]/parent::h2/parent::div/div/div/div/div[2]/div[2]/div/div[2]/div[1]/table/tbody/tr")
                            FooterCount = str(len(FooterCount))
                # -------------------Home Inspection Click------------------------------
                ElementVerify = "Inspection link text click for " + ElementVerify
                PageName = "Home"
                MdataSheetTab = "test_Smoke_Home_Admin"
                MdataSheetItem = "HomeInspectionClick"
                MdataSheetItem2 = "HomeInspectionClickText"
                ElementExpected = "Inspection Date"
                ElementActionCls.ElementActionMeth(driver, MdataSheetTab, MdataSheetItem, MdataSheetItem2,
                                                   ElementExpected,
                                                   ElementVerify,
                                                   PageName, TestResult, TestResultStatus)
                driver.find_element(By.XPATH, "//a/div[contains(text(),'Home')]").click()
                LoaderCls.LoaderMeth(driver)
            else:
                print("No Data available **************** for " + ElementVerify)
                FooterCount="0"
            # ---------------Matching Count for Finalization Required ---------------
            ElementVerify = "Finalization Required Count"
            print("FooterCount is " + FooterCount)
            ElementFound = FooterCount
            if Count == ElementFound:
                TestResult.append(
                    ElementVerify + " (" + Count + ") " + " at " + PageName + " was matching with number of records" + " (" + ElementFound + ")")
                TestResultStatus.append("Pass")
            else:
                TestResult.append(
                    ElementVerify + " (" + Count + ") " + " at " + PageName + "was not matching with number of "
                                                                                        "records" + " (" +
                    ElementFound + ")")
                TestResultStatus.append("Fail")

            #driver.find_element(By.XPATH, "//a/div[contains(text(),'Home')]").click()

        except Exception as Mainerror:
            print(Mainerror)
            stringMainerror = repr(Mainerror)
            if stringMainerror in "InvalidSessionIdException('invalid session id', None, None)":
                pass
            else:
                TestResult.append(stringMainerror)
                TestResultStatus.append("Fail")

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = DataReadMaster.GlobalData("test_Smoke_Home_Admin", "TestName")

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------